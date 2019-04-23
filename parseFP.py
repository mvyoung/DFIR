# also look into this progress bar library
# https://blog.sicara.com/perfect-python-command-line-interfaces-7d5d4efad6a2
# https://pypi.org/project/tqdm/
from progress.bar import Bar
from datetime import datetime
from collections import defaultdict
from tqdm import tqdm
import os
import time
import openpyxl
import shutil
import sys

import browser.getChromeHistory
# import getChromeHistory

# LINK
# TYPED
# AUTO_BOOKMARK
# MANUAL_SUBFRAME
# GENERATED
# START_PAGE
# FORM_SUBMIT
# KEYWORD_GENERATED

# AUTO_SUBFRAME
# RELOAD
# KEYWORD_GENERATED

malicious_categories = ['Blacklist', 'Gambling']
user_initiated_transitions = []

class FPItem(object):
   # various columns in the vuln spreadsheet
   visit_time = -1
   base_url = ""
   category = ""
   full_url = ""
   hits = -1

   # class function that initializes new FPItem object
   def __init__(self, visit_time, base_url, category, full_url, hits):
      self.visit_time = visit_time
      self.base_url = base_url
      self.category = category
      self.full_url = full_url
      self.hits = hits

   # function that turns FPItem properties into a string
   def print_FPItem(self):
      print("     Visit time: " + str(self.visit_time))
      print("     Base URL:   " + str(self.base_url))
      print("     Category:   " + str(self.category))
      print("     Full URL:   " + str(self.full_url))
      print("     Hit count:  " + str(self.hits))
      print()

# function external to the FPItem class that initializes new FPItem object
def create_FPItem(visit_time, base_url, category, full_url, hits):
   return FPItem(visit_time, base_url, category, full_url, hits)


# function opens master spreadsheet and prompts user for name of new worksheet and opens it as well
def open_files():
   # report = "browser\\investigativereports.xlsx"
   report = "investigativereports.xlsx"

   # open master vuln spreadsheet, error catching does not work because load_workbook is part of the openpyxl library and does not throw an IOError like open() does
   try:
      wb = openpyxl.load_workbook(report, read_only=False)
   except IOError:
      print("FP history file could not be opened. Run this script in the same directory as the file called 'investigativereports.xlsx'")
      return False, False

   wb_sheet = wb.worksheets[0]

   return wb, wb_sheet

# function returns list of FPItmes
def parse_FPItems(sheet):
   history = {}
   count = 0

   # bar = Bar('Reading in FP history...                              ', fill='=', max=sheet.max_row-4)
   for row in tqdm(sheet.iter_rows(row_offset=2)):
      ndx = 0
      user = ""
      datebuilder = -1
      timebuilder = -1
      datetime_obj = -1
      base_url = ""
      category = ""
      full_url = ""
      hits = -1

      # iterate through all cells in a row to populate all the variables needed for a new FPItem instance
      for cell in row:
         # column A
         if ndx == 0:
            user = cell.value
         # column B
         if ndx == 1:
            datebuilder = cell.value
         # column C
         if ndx == 2:
            timebuilder = cell.value
         # column D
         if ndx == 3:
            base_url = cell.value
         # column E
         if ndx == 4:
            category = cell.value
         # column F
         if ndx == 5:
            full_url = cell.value
         # column G
         if ndx == 6:
            hits = cell.value
         ndx += 1

      try:
         datetime_obj = datetime.strptime(str(datebuilder) + " " + str(timebuilder), "%Y-%m-%d %H:%M:%S")
         history_obj = create_FPItem(datetime_obj, base_url, category, full_url, hits)
         history[int(history_obj.visit_time.timestamp())] = history_obj
      except:
         continue
      # bar.next()
   # bar.finish()

   return history

# matches up chrome_history and fp_history by time
def match_up(chrome_history, fp_history):
   matched_items = defaultdict(list)

   # bar = Bar('Matching up Chrome history to FP history...', fill='=', max=(len(chrome_history) + len(fp_history)))
   for item in tqdm(fp_history):
      matched_items[int(fp_history[item].visit_time.timestamp())].append(fp_history[item])
      # bar.next()

   for item in tqdm(chrome_history):
      matched_items[int(chrome_history[item].visit_time.timestamp())].append(chrome_history[item])
      # bar.next()
   # bar.finish()

   return matched_items


# returns items from history that match malicious categories defined in malicious_categories
def juxtapose(history):
   malicious_items = {}

   # bar = Bar('Identifying blocked pages of interest in FP History...', fill='=', max=len(history))
   for item in tqdm(history):
      if any(malicious in history[item].category for malicious in malicious_categories):
         malicious_items[int(history[item].visit_time.timestamp())] = history[item]
      # bar.next()
   # bar.finish()

   return malicious_items

def get_chrome_history(username):
   query = getChromeHistory.db_fetch(username)
   if query:
      chrome_history = getChromeHistory.pull_history(query)
   else:
      return False

   return chrome_history

def run(opt, username):
   wb, wb_sheet = open_files()
   if wb and wb_sheet:
      fp_history = parse_FPItems(wb_sheet)
      # malicious_history = juxtapose(fp_history)

      if opt == "chrome":
         chrome_history = get_chrome_history(username)
      elif opt == "IE":
         return
      else:
         return

      timeline = match_up(chrome_history, fp_history)

      for time in tqdm(timeline):
         # print("History events for " + datetime.fromtimestamp(time).strftime("%A, %B %d, %Y %I:%M:%S"))
         for event in timeline[time]:
            if isinstance(event, FPItem):
               # event.print_FPItem()
               pass
            elif isinstance(event, getChromeHistory.HistoryItem):
               # event.print_HistoryItem()
               pass

      # if chrome_history:
      #    for key in (malicious_history.keys() | chrome_history.keys()):
      #       if key in malicious_history and key in chrome_history:
      #          print("FP URL:            " + malicious_history[key].full_url)
      #          print("Category:          " + malicious_history[key].category)
      #          print("Chrome URL:        " + chrome_history[key].url)
      #          print("Visited:           " + str(chrome_history[key].visit_time))
      #          print("How user got here: " + " - ".join(browser.getChromeHistory.page_transitions[chrome_history[key].transition]))
      #          print()
   else:
      return

def main():
   print("parseFP.py is only programmed to pull Chrome history when run independently.")
   username = input("Which user's Chrome history should be returned? ")

   if not username:
      username = os.getlogin()
   # else:
   #    username = username[0]
   run("chrome", username)

if __name__ == '__main__':
   main()